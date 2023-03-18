# result processing automation tools
 
import os
import os.path as pt
import sys
import pprint
import glob
import shutil

import openpyxl as xl
from PIL import Image
from openpyxl.drawing.image import Image as xlImage

from core.settings import BASE_DIR




PATH_TO_EXCELL_TEMPLATE = BASE_DIR / 'custom_admin/views/_Template.xlsx'
PATH_TO_STUDENT_EXCELL_FILES = BASE_DIR / 'media\media\excell_spreedsheets'      # !path to specific to windows
PATH_TO_RESIZED_PHOTO = BASE_DIR / 'custom_admin/views/resized_photo'

def make_student_report(report_student) -> str:
    st = get_student_detials(report_student)
    create_excel_file(st)
    return f'/media/media/excell_spreedsheets/{st["output_sheet_name"]}'


def get_student_detials(report_student):
    photo_path = report_student.student.headshot.path
    student_name = report_student.student.full_name()
    student_id = report_student.student.id
    return {
        'photo_path': photo_path,
        'student_name': student_name,
        'resized_photo': resize(photo_path, PATH_TO_RESIZED_PHOTO.absolute()),
        'temp': PATH_TO_EXCELL_TEMPLATE,
        'output_sheet': pt.join(PATH_TO_STUDENT_EXCELL_FILES.absolute(), f'{student_name}_{student_id}.xlsx'),
        'output_sheet_name': f'{student_name}_{student_id}.xlsx',
        'student': report_student,
        'report': report_student.rep_klass.report,
    }


def create_excel_file(student_details):
    _s = student_details
    template, student_name, resized_photo, output_sheet, student, report = (_s['temp'], _s['student_name'],
                                                            _s['resized_photo'], _s['output_sheet'], _s['student'], _s['report'])
    rep_s, student = student, student.student

    # this fills in student details
    temp_wb = xl.load_workbook(template)
    ws = temp_wb['GRADER']
    ws['B6'].value = student_name
    ws['D6'].value = str(student.klass)
    ws['B8'].value = student.admission_no
    ws['D7'].value = report.term_full()
    ws['B7'].value = report.year
    ws['D8'].value = student.serial_no

    # this fills in the scores
    cell_start = 12
    test = []
    exam = []
    test_exam = []
    for i, subject in enumerate(rep_s.reportsubject_set.all()):
        cs = str(cell_start + i)
        ws['A' + cs] = i + 1          # s/n
        ws['B' + cs] = str(subject.subject)         # su
        ws['C' + cs] = str(subject.test)          # 
        ws['D' + cs] = str(subject.exam)          #
        ws['E' + cs] = int(subject.exam) + int(subject.test)          #
        ws['F' + cs] = grade(int(subject.exam) + int(subject.test))          #

        test.append(subject.test)
        exam.append(subject.exam)
        test_exam.append(int(subject.exam) + int(subject.test))


    cell_start = (cell_start + i) + 2
    # total
    ws[f'D{cell_start}'] = 'Total :'
    ws[f'E{cell_start}'] = sum(test_exam)

    # average
    ws[f'D{cell_start+1}'] = 'Ave :'
    ws[f'E{cell_start+1}'] = round(sum(test_exam) / len(test_exam), 2)


    img_xl = xlImage(resized_photo)
    ws.add_image(img_xl, 'f5')
    temp_wb.save(output_sheet)


def resize(fname, to):
    img = Image.open(fname)
    img = img.resize((int(0.76 * 96), int(1.06 * 96)))
    path, name = pt.split(fname)
    s, ext = pt.splitext(name)
    new_path = pt.join(to, s + '_resized' + ext)

    img.save(new_path)
    return new_path



def grade(mark):
    gd = {
        'A': range(90, 101),
        'B': range(70, 90),
        'C': range(50, 70),
        'D': range(40, 50),
        'F': range(0, 40),
    }

    for m, r in gd.items():
        if mark in r:
            return m